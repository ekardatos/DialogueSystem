# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions
#
#
# This is a simple example for a custom action which utters "Hello World!"
#
#from typing import Any, Text, Dict, List
#
#from rasa_sdk import Action, Tracker
#from rasa_sdk.executor import CollectingDispatcher
#
#
#class ActionHelloWorld(Action):
#
#    def name(self) -> Text:
#        return "store_customer_action"
#        
#         #return "action_hello_world"
#
#    def run(self, dispatcher: CollectingDispatcher,
#            tracker: Tracker,
#            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#        dispatcher.utter_message(text="Hello World!")
#
#        return []
#
################################################################################

import openpyxl
from rasa_sdk import Action, Tracker
from typing import Any, Text, Dict, List
from rasa_sdk.executor import CollectingDispatcher


class StoreClientInfo(Action):
    def name(self) -> Text:
        return "store_customer_action"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        client_name = tracker.get_slot("names")
        

        # Open the excel file
        wb = openpyxl.load_workbook('C:/Users/Ευάγγελος Καρδάτος/Desktop/Rasa_actions.xlsx')
        sheet = wb['Sheet1']

        # Append the client information to the next empty row
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = client_name
        

        # Save the excel file
        wb.save('C:/Users/Ευάγγελος Καρδάτος/Desktop/Rasa_actions.xls')

        # Respond to the user
        dispatcher.utter_message(template="utter_client_info_stored")

        return []